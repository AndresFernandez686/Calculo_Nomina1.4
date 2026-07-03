"""
Calculadora de Sueldos - Aplicación Flask.

Migración completa desde Streamlit a Flask manteniendo toda la lógica de negocio:
- Carga de archivos Excel o PDF (hasta 2 quincenas).
- Configuración de valor por hora y hasta 3 feriados (doble pago).
- Detección y corrección de registros incompletos y horarios ambiguos.
- Cálculo de horas normales / especiales (20:00-22:00 +30%) y feriados (x2).
- Resultados con resumen, descarga en Excel e historial persistido en BD.
"""
import io
import os
import pickle
import re
import uuid
import calendar
from datetime import date, datetime, timedelta
from typing import Optional, cast

import pandas as pd
from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import desc, inspect, text

from payroll.data_processor import procesar_datos_excel, validar_archivo_excel
from payroll.models import (
    CalculationRun,
    Employee,
    EmployeeAttendance,
    EmployeePayroll,
    EmployeeRecord,
    HistorialSalarios,
    Liquidacion,
    PromedioLaboral,
    db,
)
from payroll.models_liquidacion import obtener_mensaje_liquidacion
from payroll.models_liquidacion import calcular_vacaciones_desde_historial
from payroll.pdf_processor import (
    convertir_a_dataframe_estandar,  # noqa: F401  (import keeps API discoverable)
    detectar_horarios_ambiguos,
    detectar_registros_incompletos,
    filtrar_registros_sin_asistencia,
    procesar_pdf_a_dataframe,
    validar_datos_pdf,
)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
PENDING_DIR = os.path.join(INSTANCE_DIR, "pending")
os.makedirs(PENDING_DIR, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "calculadora-sueldos-dev-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(
    INSTANCE_DIR, "sueldos.db"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25 MB

db.init_app(app)

from payroll.calculations import calcular_horas_especiales, horas_a_horasminutos


@app.template_filter("hhmm")
def _hhmm_filter(value):
    """Convierte horas decimales a formato H:MM para las plantillas."""
    try:
        return horas_a_horasminutos(float(value))
    except (TypeError, ValueError):
        return "0:00"


VALOR_POR_HORA_DEFAULT = 13937.0


# --------------------------------------------------------------------------- #
# Utilidades                                                                   #
# --------------------------------------------------------------------------- #
def _parse_feriados(form):
    """Lee hasta 3 fechas de feriado del formulario y devuelve un set de dates."""
    fechas = set()
    for key in ("feriado_1", "feriado_2", "feriado_3"):
        valor = (form.get(key) or "").strip()
        if valor:
            try:
                fechas.add(datetime.strptime(valor, "%Y-%m-%d").date())
            except ValueError:
                pass
    return fechas


def _normalizar_hora_pdf(valor):
    """Normaliza horas de PDF a HH:MM para evitar descarte por formato."""
    if pd.isna(valor):
        return valor

    texto = str(valor).strip()
    if not texto:
        return texto

    match = re.search(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", texto)
    if match:
        hora = int(match.group(1))
        minuto = int(match.group(2))
        if 0 <= hora <= 23 and 0 <= minuto <= 59:
            return f"{hora:02d}:{minuto:02d}"

    return texto


def _guardar_pendiente(df, config):
    """Persiste el DataFrame y la configuración en disco hasta confirmar correcciones."""
    token = uuid.uuid4().hex
    ruta = os.path.join(PENDING_DIR, f"{token}.pkl")
    with open(ruta, "wb") as fh:
        pickle.dump({"df": df, "config": config}, fh)
    return token


def _cargar_pendiente(token):
    ruta = os.path.join(PENDING_DIR, f"{token}.pkl")
    if not token or not os.path.exists(ruta):
        return None
    with open(ruta, "rb") as fh:
        return pickle.load(fh)


def _eliminar_pendiente(token):
    ruta = os.path.join(PENDING_DIR, f"{token}.pkl")
    if token and os.path.exists(ruta):
        os.remove(ruta)


def _df_incompletos_a_lista(df_incompletos):
    items = []
    for idx, row in df_incompletos.iterrows():
        items.append(
            {
                "idx": int(idx),
                "empleado": row["Empleado"],
                "fecha": _fecha_str(row["Fecha"]),
                "horario": str(row.get("Horario_Registrado", "")),
                "dato_faltante": row.get("Dato_Faltante", ""),
                "problema": row.get("Tipo_Problema", ""),
            }
        )
    return items


def _df_ambiguos_a_lista(df_ambiguos):
    items = []
    for idx, row in df_ambiguos.iterrows():
        items.append(
            {
                "idx": int(idx),
                "empleado": row["Empleado"],
                "fecha": _fecha_str(row["Fecha"]),
                "entrada": row["Entrada_Original"],
                "salida": row["Salida_Original"],
                "razon": row["Razon_Sospecha"],
            }
        )
    return items


MESES_ES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def _mes_es(month: int) -> str:
    """Nombre del mes en español (1–12).

    Envuelve el acceso a ``MESES_ES`` para evitar el falso positivo de Pylance
    con Python 3.14 donde ``int`` no se reconoce como ``SupportsIndex``.
    """
    return MESES_ES[month - 1]  # type: ignore[index]


def _agrupar_registros_por_mes(registros):
    run_ids = {int(r.run_id) for r in registros if getattr(r, "run_id", None)}
    runs_by_id = {}
    if run_ids:
        runs = [db.session.get(CalculationRun, run_id) for run_id in run_ids]
        runs = [run for run in runs if run is not None]
        runs_by_id = {run.id: run for run in runs}

    meses = {}
    for registro in registros:
        fecha_str = getattr(registro, "fecha", "")
        try:
            fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        except (TypeError, ValueError):
            continue

        clave = (fecha_dt.year, fecha_dt.month)
        if clave not in meses:
            meses[clave] = {
                "year": fecha_dt.year,
                "month": fecha_dt.month,
                "mes_nombre": f"{_mes_es(fecha_dt.month)} {fecha_dt.year}",
                "registros": [],
                "dias_trabajados": 0,
                "total_horas_normales": 0.0,
                "total_horas_especiales": 0.0,
                "total_sueldos": 0.0,
                "total_salario_bruto": 0.0,
                "monto_horas_normales": 0.0,
                "monto_horas_especiales": 0.0,
                "monto_feriados": 0.0,
                "bonificacion": 0.0,
                "total_descuento_ips": 0.0,
                "total_aporte_empleador_ips": 0.0,
                "total_ips": 0.0,
                "total_salario_neto_ips": 0.0,
                "_run_ids_contados": set(),
            }

        meses[clave]["registros"].append(registro)
        horas_normales = float(registro.horas_normales or 0)
        horas_especiales = float(registro.horas_especiales or 0)

        if (horas_normales + horas_especiales) > 0:
            meses[clave]["dias_trabajados"] += 1

        meses[clave]["total_horas_normales"] += horas_normales
        meses[clave]["total_horas_especiales"] += horas_especiales
        meses[clave]["total_sueldos"] += float(registro.sueldo_final or 0)

        run_id = getattr(registro, "run_id", None)
        if run_id and run_id not in meses[clave]["_run_ids_contados"]:
            run = runs_by_id.get(int(run_id))
            if run:
                bruto = float(run.total_salario_bruto or 0)
                if bruto <= 0:
                    bruto = float(run.total_sueldos or 0)

                monto_normal = float(
                    run.total_monto_horas_normales
                    or (float(run.total_horas_normales or 0) * float(run.valor_por_hora or 0))
                )
                monto_especial = float(
                    run.total_monto_horas_especiales
                    or (
                        float(run.total_horas_especiales or 0)
                        * float(run.valor_por_hora or 0)
                        * 1.3
                    )
                )
                monto_feriados = float(
                    run.total_monto_feriados
                    if run.total_monto_feriados is not None
                    else max(0.0, bruto - monto_normal - monto_especial)
                )
                bonificacion = float(run.total_bonificacion or 0)

                if (not run.feriados and monto_feriados <= 0.01) or abs(monto_feriados) < 0.015:
                    monto_feriados = 0.0

                descuento_ips = float(run.total_descuento_ips or 0)
                aporte_empleador = float(run.total_aporte_empleador_ips or 0)
                total_ips = float(run.total_ips or (descuento_ips + aporte_empleador))
                neto = float(run.total_salario_neto_ips or (bruto - descuento_ips))

                meses[clave]["total_salario_bruto"] += bruto
                meses[clave]["monto_horas_normales"] += monto_normal
                meses[clave]["monto_horas_especiales"] += monto_especial
                meses[clave]["monto_feriados"] += monto_feriados
                meses[clave]["bonificacion"] += bonificacion
                meses[clave]["total_descuento_ips"] += descuento_ips
                meses[clave]["total_aporte_empleador_ips"] += aporte_empleador
                meses[clave]["total_ips"] += total_ips
                meses[clave]["total_salario_neto_ips"] += neto

            meses[clave]["_run_ids_contados"].add(run_id)

    salida = []
    for clave in sorted(meses.keys()):
        mes = meses[clave]
        if abs(mes["monto_feriados"]) < 0.015:
            mes["monto_feriados"] = 0.0
        if mes["total_salario_bruto"] <= 0:
            mes["total_salario_bruto"] = mes["total_sueldos"]
        if mes["total_salario_neto_ips"] <= 0:
            mes["total_salario_neto_ips"] = mes["total_salario_bruto"]
        mes.pop("_run_ids_contados", None)
        salida.append(mes)

    return salida


def _agrupar_meses_por_anio(registros_por_mes):
    """Agrupa meses por anio para separar visualmente el historial cuando aplica."""
    if not registros_por_mes:
        return [], False

    meses_por_anio = {}
    for mes in registros_por_mes:
        year = mes.get("year")
        if year is None:
            continue
        if year not in meses_por_anio:
            meses_por_anio[year] = []
        meses_por_anio[year].append(mes)

    bloques = []
    for year in sorted(meses_por_anio.keys(), reverse=True):
        meses_ordenados = sorted(
            meses_por_anio[year],
            key=lambda item: (item.get("month") or 0),
        )
        bloques.append({"year": year, "meses": meses_ordenados})

    return bloques, len(bloques) > 1


def _filtrar_registros_por_mes(registros, year, month):
    filtrados = []
    for registro in registros:
        fecha_str = getattr(registro, "fecha", "")
        try:
            fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        except (TypeError, ValueError):
            continue
        if fecha_dt.year == year and fecha_dt.month == month:
            filtrados.append(registro)
    return filtrados


def _fecha_str(valor):
    try:
        return pd.to_datetime(valor).strftime("%d/%m/%Y")
    except Exception:  # noqa: BLE001
        return str(valor)


def _month_key_from_value(valor: object) -> Optional[str]:
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return f"{valor.year:04d}-{valor.month:02d}"

    texto = str(valor).strip()
    if not texto:
        return None

    # Soporta fechas normalizadas del sistema (YYYY-MM-DD) sin ambiguedad.
    try:
        fecha_iso = datetime.strptime(texto[:10], "%Y-%m-%d")
        return f"{fecha_iso.year:04d}-{fecha_iso.month:02d}"
    except ValueError:
        pass

    # Soporta fechas de formularios/reportes en formato DD/MM/YYYY.
    try:
        fecha_latam = datetime.strptime(texto[:10], "%d/%m/%Y")
        return f"{fecha_latam.year:04d}-{fecha_latam.month:02d}"
    except ValueError:
        pass

    try:
        fecha = pd.to_datetime(texto, errors="coerce")
    except Exception:  # noqa: BLE001
        return None
    if pd.isna(fecha):
        return None
    return f"{fecha.year:04d}-{fecha.month:02d}"


def _month_label_from_key(month_key: str) -> str:
    try:
        year_str, month_str = month_key.split("-", 1)
        year = int(year_str)
        month = int(month_str)
        return f"{_mes_es(month)} {year}"
    except (ValueError, IndexError):
        return month_key


def _obtener_meses_df(df) -> set[str]:
    if "Fecha" not in df.columns:
        return set()
    meses: set[str] = set()
    for valor in df["Fecha"].tolist():
        month_key = _month_key_from_value(valor)
        if month_key:
            meses.add(month_key)
    return meses


def _agrupar_registros_por_mes_key(registros):
    meses = {}
    for registro in registros:
        month_key = _month_key_from_value(getattr(registro, "fecha", ""))
        if not month_key:
            continue

        if month_key not in meses:
            meses[month_key] = {
                "month_key": month_key,
                "mes_nombre": _month_label_from_key(month_key),
                "registros": [],
                "dias_trabajados": 0,
                "total_horas_normales": 0.0,
                "total_horas_especiales": 0.0,
                "total_sueldos": 0.0,
            }

        meses[month_key]["registros"].append(registro)
        horas_normales = float(getattr(registro, "horas_normales", 0) or 0)
        horas_especiales = float(getattr(registro, "horas_especiales", 0) or 0)
        if (horas_normales + horas_especiales) > 0:
            meses[month_key]["dias_trabajados"] += 1
        meses[month_key]["total_horas_normales"] += horas_normales
        meses[month_key]["total_horas_especiales"] += horas_especiales
        meses[month_key]["total_sueldos"] += float(getattr(registro, "sueldo_final", 0) or 0)

    return [meses[key] for key in sorted(meses.keys(), reverse=True)]


def _obtener_conflictos_mes_empleado(employee_id: int, meses_objetivo: set[str]):
    if not meses_objetivo:
        return [], []

    registros = (
        EmployeePayroll.query.filter_by(employee_id=employee_id)
        .order_by(desc(EmployeePayroll.created_at))
        .all()
    )
    registros_conflictivos = []
    for registro in registros:
        month_key = _month_key_from_value(getattr(registro, "fecha", ""))
        if month_key in meses_objetivo:
            registros_conflictivos.append(registro)

    meses_conflictivos_set: set[str] = set()
    for registro in registros_conflictivos:
        month_key = _month_key_from_value(getattr(registro, "fecha", ""))
        if month_key:
            meses_conflictivos_set.add(month_key)

    meses_conflictivos = sorted(meses_conflictivos_set, reverse=True)
    return meses_conflictivos, registros_conflictivos


def _reemplazar_meses_existentes(employee_id: int, meses_objetivo: set[str]) -> None:
    if not meses_objetivo:
        return

    registros = EmployeePayroll.query.filter_by(employee_id=employee_id).all()
    registros_a_borrar = []
    run_ids: set[int] = set()
    for registro in registros:
        month_key = _month_key_from_value(getattr(registro, "fecha", ""))
        if month_key in meses_objetivo:
            registros_a_borrar.append(registro)
            if getattr(registro, "run_id", None):
                run_ids.add(int(registro.run_id))

    for registro in registros_a_borrar:
        db.session.delete(registro)

    if run_ids:
        for run_id in run_ids:
            EmployeeAttendance.query.filter_by(run_id=run_id).delete(
                synchronize_session=False
            )
            EmployeeRecord.query.filter_by(run_id=run_id).delete(
                synchronize_session=False
            )
            CalculationRun.query.filter_by(id=run_id).delete(
                synchronize_session=False
            )

    db.session.commit()


def _sanear_nombre_para_archivo(nombre: str) -> str:
    texto = re.sub(r"[^\w\-]+", "_", nombre.strip(), flags=re.UNICODE)
    texto = re.sub(r"_+", "_", texto).strip("_")
    return texto or "empleado"


def _generar_df_desde_registros(registros):
    filas = []
    for r in registros:
        filas.append(
            {
                "Fecha": r.fecha,
                "Entrada": r.entrada,
                "Salida": r.salida,
                "Feriado": r.feriado,
                "Horas Trabajadas (h:mm)": r.horas_trabajadas,
                "Horas Normales": horas_a_horasminutos(r.horas_normales),
                "Horas Especiales": horas_a_horasminutos(r.horas_especiales),
                "Descuento Inventario": r.descuento_inventario,
                "Descuento Caja": r.descuento_caja,
                "Retiro": r.retiro,
                "Sueldo Final": r.sueldo_final,
            }
        )
    return pd.DataFrame(filas)


def _crear_pdf_desde_registros(registros, empleado_nombre, mes_nombre):
    df = _generar_df_desde_registros(registros)
    resumen = {
        "total_horas_normales": sum(float(r.horas_normales or 0) for r in registros),
        "total_horas_especiales": sum(float(r.horas_especiales or 0) for r in registros),
        "total_sueldos": sum(float(r.sueldo_final or 0) for r in registros),
    }

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"{empleado_nombre} — {mes_nombre}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    columnas = [
        "Fecha",
        "Entrada",
        "Salida",
        "Feriado",
        "Horas Normales",
        "Horas Especiales",
        "Desc. Inv.",
        "Desc. Caja",
        "Retiro",
        "Sueldo Final",
    ]
    datos = [columnas]
    for _, fila in df.iterrows():
        datos.append(
            [
                fila["Fecha"],
                fila["Entrada"] or "-",
                fila["Salida"] or "-",
                fila["Feriado"] or "-",
                fila["Horas Normales"],
                fila["Horas Especiales"],
                f"${fila['Descuento Inventario']:.0f}" if fila["Descuento Inventario"] else "-",
                f"${fila['Descuento Caja']:.0f}" if fila["Descuento Caja"] else "-",
                f"${fila['Retiro']:.0f}" if fila["Retiro"] else "-",
                f"${fila['Sueldo Final']:.0f}",
            ]
        )

    tabla = Table(datos, repeatRows=1, hAlign="LEFT")
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#3a485d")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7d7d7")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7d7d7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(tabla)
    story.append(Spacer(1, 16))

    resumen_tabla = Table(
        [
            ["Total Horas Normales", f"{resumen['total_horas_normales']:.2f}"],
            ["Total Horas Especiales", f"{resumen['total_horas_especiales']:.2f}"],
            ["Sueldo Total", f"${resumen['total_sueldos']:.0f}"],
        ],
        hAlign="LEFT",
        colWidths=[150, 120],
    )
    resumen_tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f9f9f9")),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7d7d7")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7d7d7")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(resumen_tabla)

    doc.build(story)
    output.seek(0)
    return output


def _crear_excel_desde_registros(registros, empleado_nombre, mes_nombre):
    df = _generar_df_desde_registros(registros)
    resumen = {
        "total_horas_normales": sum(float(r.horas_normales or 0) for r in registros),
        "total_horas_especiales": sum(float(r.horas_especiales or 0) for r in registros),
        "total_sueldos": sum(float(r.sueldo_final or 0) for r in registros),
    }
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sueldos")
        workbook = writer.book
        worksheet = writer.sheets["Sueldos"]

        fila_inicio = len(df) + 3
        bold = Font(bold=True)
        worksheet.cell(row=fila_inicio, column=1, value="Total Horas Normales").font = bold
        worksheet.cell(row=fila_inicio, column=2, value=resumen["total_horas_normales"]).font = bold
        worksheet.cell(row=fila_inicio + 1, column=1, value="Total Horas Especiales").font = bold
        worksheet.cell(row=fila_inicio + 1, column=2, value=resumen["total_horas_especiales"]).font = bold
        worksheet.cell(row=fila_inicio + 2, column=1, value="Sueldo Total").font = bold
        worksheet.cell(row=fila_inicio + 2, column=2, value=resumen["total_sueldos"]).font = bold

    output.seek(0)
    return output


def _hhmm_to_decimal(valor):
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        try:
            return float(valor)
        except (TypeError, ValueError):
            return 0.0
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "nat", "none"}:
        return 0.0
    if ":" in texto:
        h_str, m_str = texto.split(":", 1)
        try:
            horas = float(h_str)
            minutos = float(m_str)
            return horas + minutos / 60.0
        except ValueError:
            return 0.0
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _parsear_feriados_run(run: CalculationRun):
    fechas = set()
    if not run.feriados:
        return fechas
    for valor in str(run.feriados).split(","):
        texto = valor.strip()
        if not texto:
            continue
        try:
            fechas.add(datetime.strptime(texto, "%d/%m/%Y").date())
        except ValueError:
            continue
    return fechas


def _registro_tiene_horario_valido(record: EmployeeRecord) -> bool:
    entrada = str(record.entrada or "").strip()
    salida = str(record.salida or "").strip()
    invalidos = {"", "nan", "0:00", "00:00", "none"}
    return entrada.lower() not in invalidos and salida.lower() not in invalidos


def _run_necesita_recalculo(run: CalculationRun, records: list[EmployeeRecord]) -> bool:
    for record in records:
        if not _registro_tiene_horario_valido(record):
            continue
        observaciones = str(record.observaciones or "").strip().lower()
        if "entrada ajustada al inicio de la jornada laboral" in observaciones:
            return True
        if "salida ajustada al fin de la jornada laboral" in observaciones:
            return True
        if str(record.horas_trabajadas or "").strip() == "0:00" and record.sueldo_final == 0:
            return True
    return False


def _recalcular_run_legado(run: CalculationRun) -> None:
    records = cast(list[EmployeeRecord], list(getattr(run, "records", []) or []))
    if not records or not _run_necesita_recalculo(run, records):
        return

    filas = []
    for record in records:
        filas.append(
            {
                "Empleado": record.empleado,
                "Fecha": record.fecha,
                "Entrada": record.entrada,
                "Salida": record.salida,
                "Descuento Inventario": record.descuento_inventario,
                "Descuento Caja": record.descuento_caja,
                "Retiro": record.retiro,
            }
        )

    resultado = procesar_datos_excel(
        pd.DataFrame(filas),
        run.valor_por_hora,
        _parsear_feriados_run(run),
    )
    if len(resultado["resultados"]) != len(records):
        return

    payroll_rows = cast(
        list[EmployeePayroll],
        list(EmployeePayroll.query.filter_by(run_id=run.id).all()),
    )
    payroll_rows.sort(key=lambda row: row.id)

    for record, fila in zip(records, resultado["resultados"]):
        record.feriado = fila["Feriado"]
        record.horas_trabajadas = fila["Horas Trabajadas (h:mm)"]
        record.horas_normales = _hhmm_to_decimal(fila["Horas Normales"])
        record.horas_especiales = _hhmm_to_decimal(fila["Horas Especiales"])
        record.monto_horas_normales = float(fila.get("Monto Horas Normales") or 0)
        record.monto_horas_especiales = float(fila.get("Monto Horas Especiales") or 0)
        record.monto_feriado = float(fila.get("Monto Feriado") or 0)
        record.bonificacion = float(fila.get("Bonificacion") or 0)
        record.sueldo_bruto = float(fila.get("Sueldo Bruto") or 0)
        record.descuento_ips = float(fila.get("Descuento IPS") or 0)
        record.sueldo_final = float(fila["Sueldo Final"] or 0)
        record.observaciones = fila.get("Observaciones", "")

    for payroll, fila in zip(payroll_rows, resultado["resultados"]):
        payroll.feriado = fila["Feriado"]
        payroll.horas_trabajadas = fila["Horas Trabajadas (h:mm)"]
        payroll.horas_normales = _hhmm_to_decimal(fila["Horas Normales"])
        payroll.horas_especiales = _hhmm_to_decimal(fila["Horas Especiales"])
        payroll.monto_horas_normales = float(fila.get("Monto Horas Normales") or 0)
        payroll.monto_horas_especiales = float(fila.get("Monto Horas Especiales") or 0)
        payroll.monto_feriado = float(fila.get("Monto Feriado") or 0)
        payroll.bonificacion = float(fila.get("Bonificacion") or 0)
        payroll.sueldo_bruto = float(fila.get("Sueldo Bruto") or 0)
        payroll.descuento_ips = float(fila.get("Descuento IPS") or 0)
        payroll.sueldo_final = float(fila["Sueldo Final"] or 0)
        payroll.observaciones = fila.get("Observaciones", "")

    run.total_horas = resultado["total_horas"]
    run.total_horas_normales = resultado["total_horas_normales"]
    run.total_horas_especiales = resultado["total_horas_especiales"]
    run.total_monto_horas_normales = resultado["total_monto_horas_normales"]
    run.total_monto_horas_especiales = resultado["total_monto_horas_especiales"]
    run.total_monto_feriados = resultado["total_monto_feriados"]
    run.total_bonificacion = resultado["total_bonificacion"]
    run.total_sueldos = resultado["total_sueldos"]
    run.total_salario_bruto = resultado["total_salario_bruto"]
    run.total_descuento_ips = resultado["total_descuento_ips"]
    run.total_aporte_empleador_ips = resultado["total_aporte_empleador_ips"]
    run.total_ips = resultado["total_ips"]
    run.total_salario_neto_ips = resultado["total_salario_neto_ips"]
    run.total_registros = len(resultado["resultados"])
    db.session.commit()


def _normalizar_nombre(nombre: str) -> str:
    texto = str(nombre or "").strip().lower()
    texto = texto.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    texto = re.sub(r"[^a-z0-9\s]", "", texto)
    return " ".join(texto.split())


def _empleado_coincide(nombre_extraido: str, empleado: Employee) -> bool:
    nombre_extraido = _normalizar_nombre(nombre_extraido)
    nombre_empleado = _normalizar_nombre(empleado.nombre)
    if not nombre_extraido or not nombre_empleado:
        return False
    if nombre_extraido == nombre_empleado:
        return True

    tokens_extraido = nombre_extraido.split()
    tokens_empleado = nombre_empleado.split()

    if len(tokens_extraido) == 1 and tokens_empleado:
        return next(iter(tokens_extraido)) == next(iter(tokens_empleado))
    if len(tokens_empleado) == 1 and tokens_extraido:
        return next(iter(tokens_empleado)) == next(iter(tokens_extraido))

    return _tokens_consecutivos(tokens_extraido, tokens_empleado) or _tokens_consecutivos(
        tokens_empleado,
        tokens_extraido,
    )


def _tokens_consecutivos(tokens_base: list[str], tokens_objetivo: list[str]) -> bool:
    if len(tokens_objetivo) > len(tokens_base):
        return False
    for indice in range(len(tokens_base) - len(tokens_objetivo) + 1):
        if tokens_base[indice: indice + len(tokens_objetivo)] == tokens_objetivo:
            return True
    return False


def _desglose_horas(entrada_texto: str, salida_texto: str) -> dict:
    entrada_dt = datetime.strptime(entrada_texto, "%H:%M")
    salida_dt = datetime.strptime(salida_texto, "%H:%M")
    if salida_dt < entrada_dt:
        salida_dt += timedelta(days=1)

    horas_totales = (salida_dt - entrada_dt).total_seconds() / 3600.0
    horas_normales, horas_especiales = calcular_horas_especiales(entrada_dt, salida_dt)

    return {
        "entrada": entrada_texto,
        "salida": salida_texto,
        "horas_totales_decimal": round(horas_totales, 10),
        "horas_totales_hhmm": horas_a_horasminutos(horas_totales),
        "horas_normales_decimal": round(horas_normales, 10),
        "horas_normales_hhmm": horas_a_horasminutos(horas_normales),
        "horas_especiales_decimal": round(horas_especiales, 10),
        "horas_especiales_hhmm": horas_a_horasminutos(horas_especiales),
    }


def _calcular_y_guardar(df, config, employee: Employee):
    """Ejecuta el cálculo y guarda el resultado en la base de datos."""
    resultado = procesar_datos_excel(
        df,
        config["valor_por_hora"],
        set(config["feriados"]),
        ips_enabled=config.get("ips_enabled", False),
    )

    run = CalculationRun()
    run.employee_id = employee.id
    run.source_name = config.get("source_name")
    run.source_type = config.get("source_type", "excel")
    run.valor_por_hora = config["valor_por_hora"]
    run.feriados = (
        ", ".join(d.strftime("%d/%m/%Y") for d in sorted(config["feriados"]))
        if config["feriados"]
        else None
    )
    run.seguro_ips = "Sí" if config.get("ips_enabled", False) else "No"
    run.total_horas = resultado["total_horas"]
    run.total_horas_normales = resultado["total_horas_normales"]
    run.total_horas_especiales = resultado["total_horas_especiales"]
    run.total_monto_horas_normales = resultado["total_monto_horas_normales"]
    run.total_monto_horas_especiales = resultado["total_monto_horas_especiales"]
    run.total_monto_feriados = resultado["total_monto_feriados"]
    run.total_bonificacion = resultado["total_bonificacion"]
    run.total_sueldos = resultado["total_sueldos"]
    run.total_salario_bruto = resultado["total_salario_bruto"]
    run.total_descuento_ips = resultado["total_descuento_ips"]
    run.total_aporte_empleador_ips = resultado["total_aporte_empleador_ips"]
    run.total_ips = resultado["total_ips"]
    run.total_salario_neto_ips = resultado["total_salario_neto_ips"]
    run.total_registros = len(resultado["resultados"])

    db.session.add(run)
    db.session.flush()

    for fila in resultado["resultados"]:
        record = EmployeeRecord()
        record.run_id = run.id
        record.employee_id = employee.id
        record.empleado = fila["Empleado"]
        record.fecha = fila["Fecha"]
        record.entrada = fila["Entrada"]
        record.salida = fila["Salida"]
        record.feriado = fila["Feriado"]
        record.horas_trabajadas = fila["Horas Trabajadas (h:mm)"]
        record.horas_normales = _hhmm_to_decimal(fila["Horas Normales"])
        record.horas_especiales = _hhmm_to_decimal(fila["Horas Especiales"])
        record.monto_horas_normales = float(fila.get("Monto Horas Normales") or 0)
        record.monto_horas_especiales = float(fila.get("Monto Horas Especiales") or 0)
        record.monto_feriado = float(fila.get("Monto Feriado") or 0)
        record.bonificacion = float(fila.get("Bonificacion") or 0)
        record.sueldo_bruto = float(fila.get("Sueldo Bruto") or 0)
        record.descuento_inventario = float(fila["Descuento Inventario"] or 0)
        record.descuento_caja = float(fila["Descuento Caja"] or 0)
        record.retiro = float(fila["Retiro"] or 0)
        record.descuento_ips = float(fila["Descuento IPS"] or 0)
        record.sueldo_final = float(fila["Sueldo Final"] or 0)
        record.observaciones = fila.get("Observaciones", "")
        db.session.add(record)

        payroll = EmployeePayroll()
        payroll.employee_id = employee.id
        payroll.fecha = fila["Fecha"]
        payroll.entrada = fila["Entrada"]
        payroll.salida = fila["Salida"]
        payroll.feriado = fila["Feriado"]
        payroll.horas_trabajadas = fila["Horas Trabajadas (h:mm)"]
        payroll.horas_normales = _hhmm_to_decimal(fila["Horas Normales"])
        payroll.horas_especiales = _hhmm_to_decimal(fila["Horas Especiales"])
        payroll.monto_horas_normales = float(fila.get("Monto Horas Normales") or 0)
        payroll.monto_horas_especiales = float(fila.get("Monto Horas Especiales") or 0)
        payroll.monto_feriado = float(fila.get("Monto Feriado") or 0)
        payroll.bonificacion = float(fila.get("Bonificacion") or 0)
        payroll.sueldo_bruto = float(fila.get("Sueldo Bruto") or 0)
        payroll.descuento_inventario = float(fila["Descuento Inventario"] or 0)
        payroll.descuento_caja = float(fila["Descuento Caja"] or 0)
        payroll.descuento_ips = float(fila["Descuento IPS"] or 0)
        payroll.retiro = float(fila["Retiro"] or 0)
        payroll.sueldo_final = float(fila["Sueldo Final"] or 0)
        payroll.observaciones = fila.get("Observaciones", "")
        payroll.run_id = run.id
        payroll.valor_hora_utilizado = config["valor_por_hora"]
        db.session.add(payroll)

    attendance = EmployeeAttendance()
    attendance.employee_id = employee.id
    attendance.run_id = run.id
    attendance.source_name = run.source_name or "asistencia"
    attendance.source_type = run.source_type
    attendance.total_registros = len(resultado["resultados"])
    db.session.add(attendance)

    db.session.commit()
    return run.id, resultado


def _render_confirmacion_reemplazo(empleado: Employee, pendiente_token: str, meses_conflictivos: list[str], registros_conflictivos):
    return render_template(
        "confirm_replace_month.html",
        empleado=empleado,
        pending_token=pendiente_token,
        meses_conflictivos=meses_conflictivos,
        meses_conflictivos_label=[_month_label_from_key(m) for m in meses_conflictivos],
        registros_conflictivos_por_mes=_agrupar_registros_por_mes_key(registros_conflictivos),
    )


# --------------------------------------------------------------------------- #
# Rutas                                                                        #
# --------------------------------------------------------------------------- #
@app.route("/")
def index():
    selected_employee_id = request.args.get("empleado_id", type=int)
    empleados = Employee.query.order_by(Employee.nombre).all()
    selected_employee = None
    history_runs = []

    if selected_employee_id:
        selected_employee = db.session.get(Employee, selected_employee_id)
        if selected_employee:
            history_runs = (
                CalculationRun.query.filter_by(employee_id=selected_employee_id)
                .order_by(desc(CalculationRun.created_at))
                .limit(8)
                .all()
            )

    return render_template(
        "index.html",
        empleados=empleados,
        selected_employee=selected_employee,
        selected_employee_id=selected_employee_id,
        history_runs=history_runs,
        valor_default=VALOR_POR_HORA_DEFAULT,
        hoy=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/api/empleados", methods=["GET"])
def api_empleados():
    """Devuelve la lista de empleados en JSON."""
    empleados = Employee.query.order_by(Employee.nombre).all()
    return {
        "empleados": [{"id": e.id, "nombre": e.nombre} for e in empleados]
    }


@app.route("/api/empleados/crear", methods=["POST"])
def api_crear_empleado():
    """Crea un nuevo empleado."""
    datos = request.get_json()
    nombre = (datos.get("nombre") or "").strip()
    
    if not nombre:
        return {"success": False, "error": "El nombre es requerido"}, 400
    
    empleado_existente = Employee.query.filter_by(nombre=nombre).first()
    if empleado_existente:
        return {"success": False, "error": "El empleado ya existe"}, 400
    
    empleado = Employee(nombre=nombre)  # type: ignore[call-arg]
    db.session.add(empleado)
    db.session.commit()
    
    return {
        "success": True,
        "empleado": {"id": empleado.id, "nombre": empleado.nombre}
    }, 201


@app.route("/api/validar-horas", methods=["POST"])
def api_validar_horas():
    """Valida una resta de horas puntual usando la misma lógica del sistema."""
    datos = request.get_json(silent=True) or {}
    entrada = str(datos.get("entrada") or "").strip()
    salida = str(datos.get("salida") or "").strip()

    if not entrada or not salida:
        return {
            "success": False,
            "error": "Los campos entrada y salida son requeridos en formato HH:MM",
        }, 400

    try:
        desglose = _desglose_horas(entrada, salida)
    except ValueError:
        return {
            "success": False,
            "error": "Formato inválido. Usa HH:MM, por ejemplo 10:10 y 17:10",
        }, 400

    return {
        "success": True,
        "metodo": "resta_exacta_entrada_salida",
        "fuente_externa": {
            "url": "https://calculadorasonline.com/calculadora-de-horas-minutos-y-segundos-sumar-horas-restar-horas/",
            "api_publica_detectada": False,
            "nota": "La página pública no expone un endpoint API visible; esta respuesta replica la resta exacta usada para contrastar resultados.",
        },
        "resultado": desglose,
    }


@app.route("/empleado/<int:empleado_id>/historial-salarial", methods=["GET", "POST"])
def historial_salarial(empleado_id: int):
    """Muestra y gestiona el historial de cambios de valor por hora."""
    empleado = db.session.get(Employee, empleado_id)
    if not empleado:
        abort(404)

    if request.method == "POST":
        fecha_inicio = (request.form.get("fecha_inicio") or "").strip()
        try:
            valor_hora = float(request.form.get("valor_hora") or 0)
        except ValueError:
            valor_hora = 0.0
        motivo_cambio = (request.form.get("motivo_cambio") or "Ingreso").strip()
        observaciones = (request.form.get("observaciones") or "").strip() or None
        usuario = (request.form.get("usuario") or "").strip() or None

        if not fecha_inicio or valor_hora <= 0:
            flash("Fecha de inicio y valor por hora son obligatorios.", "error")
        else:
            # Cerrar el registro vigente anterior
            registro_vigente = (
                HistorialSalarios.query
                .filter(
                    text("empleado_id = :eid AND fecha_fin IS NULL")
                    .bindparams(eid=empleado_id)
                )
                .order_by(desc(HistorialSalarios.fecha_inicio))
                .first()
            )
            if registro_vigente and registro_vigente.fecha_inicio < fecha_inicio:
                try:
                    fi = date.fromisoformat(fecha_inicio)
                    registro_vigente.fecha_fin = (fi - timedelta(days=1)).isoformat()
                except ValueError:
                    pass

            nuevo = HistorialSalarios()  # type: ignore[call-arg]
            nuevo.empleado_id = empleado_id
            nuevo.fecha_inicio = fecha_inicio
            nuevo.fecha_fin = None
            nuevo.valor_hora = valor_hora
            nuevo.motivo_cambio = motivo_cambio
            nuevo.observaciones = observaciones
            nuevo.usuario = usuario
            db.session.add(nuevo)
            db.session.commit()
            flash(
                f"Nuevo valor por hora Gs. {valor_hora:,.0f} registrado desde {fecha_inicio}.",
                "success",
            )
        return redirect(url_for("historial_salarial", empleado_id=empleado_id))

    historial = (
        HistorialSalarios.query
        .filter_by(empleado_id=empleado_id)
        .order_by(desc(HistorialSalarios.fecha_inicio))
        .all()
    )
    valor_actual = _obtener_valor_hora_vigente(empleado_id)
    return render_template(
        "historial_salarial.html",
        empleado=empleado,
        historial=historial,
        motivos=HistorialSalarios.MOTIVOS,
        valor_actual=valor_actual,
        hoy=date.today().isoformat(),
    )


@app.route("/empleado/<int:empleado_id>")
def ver_empleado(empleado_id):
    """Muestra la nómina de un empleado."""
    empleado = db.session.get(Employee, empleado_id)
    if not empleado:
        abort(404)
    
    registros = (
        EmployeePayroll.query.filter_by(employee_id=empleado_id)
        .order_by(desc(EmployeePayroll.created_at))
        .all()
    )
    registros_por_mes = _agrupar_registros_por_mes(registros)
    bloques_anuales, mostrar_separador_anual = _agrupar_meses_por_anio(registros_por_mes)
    
    return render_template(
        "employee_payroll.html",
        empleado=empleado,
        registros=registros,
        registros_por_mes=registros_por_mes,
        bloques_anuales=bloques_anuales,
        mostrar_separador_anual=mostrar_separador_anual,
    )


def _parse_fecha_form(valor):
    texto = (valor or "").strip()
    if not texto:
        return None
    try:
        return datetime.strptime(texto, "%Y-%m-%d").date()
    except ValueError:
        return None


def _formatear_fecha_iso(valor: date | None) -> str:
    return valor.isoformat() if valor else ""


def _fmt_hhmm(h: float) -> str:
    """Convierte horas decimales a formato HH:MM (ej. 14.90 -> '14:54')."""
    hrs = int(h)
    mins = int(round((h - hrs) * 60))
    return f"{hrs}:{mins:02d}"


def _obtener_valor_hora_vigente(empleado_id: int, fecha: date | None = None) -> float:
    """Devuelve el valor por hora vigente del historial salarial para un empleado.

    Busca el registro en ``historial_salarios`` donde:
    ``fecha_inicio <= fecha`` y ``(fecha_fin IS NULL OR fecha_fin >= fecha)``.
    Si no existe historial registrado, devuelve el valor por defecto global.
    """
    if fecha is None:
        fecha = date.today()
    fecha_str = fecha.isoformat()
    historial = (
        HistorialSalarios.query
        .filter(
            text(
                "empleado_id = :eid "
                "AND fecha_inicio <= :fs "
                "AND (fecha_fin IS NULL OR fecha_fin >= :fs)"
            ).bindparams(eid=empleado_id, fs=fecha_str)
        )
        .order_by(desc(HistorialSalarios.fecha_inicio))
        .first()
    )
    return float(historial.valor_hora) if historial else VALOR_POR_HORA_DEFAULT


def _obtener_fecha_ingreso_empleado(empleado: Employee, registros: list[EmployeePayroll]) -> date | None:
    if getattr(empleado, "hire_date", None):
        fecha = _parse_fecha_form(empleado.hire_date)
        if fecha:
            return fecha

    fechas = []
    for registro in registros:
        fecha = _parse_fecha_form(getattr(registro, "fecha", None))
        if fecha:
            fechas.append(fecha)

    if not fechas:
        return None

    fecha_ingreso = min(fechas)
    empleado.hire_date = fecha_ingreso.isoformat()
    return fecha_ingreso


def _obtener_ultimos_periodos(registros: list[EmployeePayroll], fecha_limite: date | None, cantidad: int = 6) -> list[dict]:
    meses = sorted(
        _agrupar_registros_por_mes(registros),
        key=lambda mes: (int(mes.get("year") or 0), int(mes.get("month") or 0)),
    )
    meses_filtrados = []
    for mes in meses:
        year = int(mes.get("year") or 0)
        month = int(mes.get("month") or 0)
        if not year or not month:
            continue
        if fecha_limite and (year, month) > (fecha_limite.year, fecha_limite.month):
            continue
        meses_filtrados.append(mes)

    if cantidad > 0 and len(meses_filtrados) > cantidad:
        return meses_filtrados[-cantidad:]
    return meses_filtrados


def _calcular_salario_pendiente_liquidacion(
    registros: list[EmployeePayroll],
    fecha_salida: date | None,
    ultimo_sueldo_pagado: bool = False,
) -> dict:
    """Calcula el salario pendiente con desglose auditable de horas."""
    _SIN_DATOS: dict = {
        "dias": 0, "horas": 0.0, "horas_normales": 0.0, "horas_especiales": 0.0, "horas_feriado": 0.0,
        "valor_hora": VALOR_POR_HORA_DEFAULT, "monto": 0.0, "periodo": "Sin periodo",
        "fecha_inicio": "", "fecha_fin": fecha_salida.strftime("%d/%m/%Y") if fecha_salida else "",
        "modo": "sin_datos",
    }
    if not registros or not fecha_salida:
        return _SIN_DATOS

    periodo_registros = []
    for registro in registros:
        fecha = _parse_fecha_form(getattr(registro, "fecha", None))
        if not fecha:
            continue
        if fecha.year == fecha_salida.year and fecha.month == fecha_salida.month and fecha <= fecha_salida:
            periodo_registros.append(registro)

    dias_periodo = max(int(fecha_salida.day), 1)
    if not periodo_registros:
        meses_historicos = _agrupar_registros_por_mes(registros)
        if not meses_historicos:
            return _SIN_DATOS

        ultimo_mes: dict = cast(dict, next(iter(meses_historicos)))
        periodo_registros_ultimo = list(ultimo_mes.get("registros") or [])
        if not periodo_registros_ultimo:
            return _SIN_DATOS

        run_id = next((reg.run_id for reg in periodo_registros_ultimo if reg.run_id is not None), None)
        valor_hora = VALOR_POR_HORA_DEFAULT
        if run_id is not None:
            run = CalculationRun.query.get(int(run_id))
            if run and float(run.valor_por_hora or 0) > 0:
                valor_hora = float(run.valor_por_hora)

        horas_normales_mes   = sum(float(reg.horas_normales or 0)   for reg in periodo_registros_ultimo)
        horas_especiales_mes = sum(float(reg.horas_especiales or 0) for reg in periodo_registros_ultimo)
        horas_feriado_mes    = sum(float(getattr(reg, "horas_feriado", 0) or 0) for reg in periodo_registros_ultimo)
        horas_total_mes = horas_normales_mes + horas_especiales_mes + horas_feriado_mes
        monto_mes_completo = sum(float(reg.sueldo_final or 0) for reg in periodo_registros_ultimo)
        dias_registrados_mes = max(len(periodo_registros_ultimo), 1)

        fecha_ref = _parse_fecha_form(getattr(next(iter(periodo_registros_ultimo)), "fecha", None))
        mes_ref = (
            f"{_mes_es(fecha_ref.month)} {fecha_ref.year}"
            if fecha_ref and 1 <= fecha_ref.month <= 12
            else "último mes"
        )

        if not ultimo_sueldo_pagado:
            fecha_fin_mes = (
                date(fecha_ref.year, fecha_ref.month,
                     calendar.monthrange(fecha_ref.year, fecha_ref.month)[1])
                if fecha_ref else None
            )
            return {
                "dias": dias_registrados_mes,
                "horas": round(horas_total_mes, 2),
                "horas_normales":   round(horas_normales_mes, 2),
                "horas_especiales": round(horas_especiales_mes, 2),
                "horas_feriado":    round(horas_feriado_mes, 2),
                "valor_hora": round(valor_hora, 2),
                "monto": round(monto_mes_completo, 2),
                "periodo": f"Mes completo ({mes_ref})",
                "fecha_inicio": fecha_ref.strftime("%d/%m/%Y") if fecha_ref else "",
                "fecha_fin":    fecha_fin_mes.strftime("%d/%m/%Y") if fecha_fin_mes else "",
                "modo": "mes_completo",
            }
        else:
            dias_transcurridos = max((fecha_salida - date(fecha_salida.year, fecha_salida.month, 1)).days + 1, 1)
            monto_promedio_dia    = monto_mes_completo / dias_registrados_mes
            h_n_dia = horas_normales_mes   / dias_registrados_mes
            h_e_dia = horas_especiales_mes / dias_registrados_mes
            h_f_dia = horas_feriado_mes    / dias_registrados_mes
            return {
                "dias": dias_transcurridos,
                "horas": round((horas_total_mes / dias_registrados_mes) * dias_transcurridos, 2),
                "horas_normales":   round(h_n_dia * dias_transcurridos, 2),
                "horas_especiales": round(h_e_dia * dias_transcurridos, 2),
                "horas_feriado":    round(h_f_dia * dias_transcurridos, 2),
                "valor_hora": round(valor_hora, 2),
                "monto": round(monto_promedio_dia * dias_transcurridos, 2),
                "periodo": f"Días parciales en {_mes_es(fecha_salida.month)} {fecha_salida.year} (ref. {mes_ref})",
                "fecha_inicio": date(fecha_salida.year, fecha_salida.month, 1).strftime("%d/%m/%Y"),
                "fecha_fin":    fecha_salida.strftime("%d/%m/%Y"),
                "modo": "dias_parciales",
            }

    run_id = next((reg.run_id for reg in periodo_registros if reg.run_id is not None), None)
    valor_hora = VALOR_POR_HORA_DEFAULT
    if run_id is not None:
        run = CalculationRun.query.get(int(run_id))
        if run and float(run.valor_por_hora or 0) > 0:
            valor_hora = float(run.valor_por_hora)

    fecha_periodo = _parse_fecha_form(getattr(next(iter(periodo_registros)), "fecha", None))

    horas_normales   = sum(float(reg.horas_normales or 0)   for reg in periodo_registros)
    horas_especiales = sum(float(reg.horas_especiales or 0) for reg in periodo_registros)
    horas_feriado    = sum(float(getattr(reg, "horas_feriado", 0) or 0) for reg in periodo_registros)
    horas_total = horas_normales + horas_especiales + horas_feriado
    monto = sum(float(reg.sueldo_final or 0) for reg in periodo_registros)
    return {
        "dias": min(len(periodo_registros), dias_periodo),
        "horas": round(horas_total, 2),
        "horas_normales":   round(horas_normales, 2),
        "horas_especiales": round(horas_especiales, 2),
        "horas_feriado":    round(horas_feriado, 2),
        "valor_hora": round(valor_hora, 2),
        "monto": round(monto, 2),
        "periodo": f"{fecha_periodo.strftime('%d/%m/%Y') if fecha_periodo else ''} al {fecha_salida.strftime('%d/%m/%Y')}",
        "fecha_inicio": fecha_periodo.strftime("%d/%m/%Y") if fecha_periodo else "",
        "fecha_fin":    fecha_salida.strftime("%d/%m/%Y"),
        "modo": "periodo_actual",
    }


def _calcular_aguinaldo_proporcional(
    registros: list[EmployeePayroll],
    fecha_salida: date | None,
    fecha_ingreso: date | None = None,
) -> dict:
    periodos = _obtener_ultimos_periodos(registros, fecha_salida, cantidad=6)

    # Mapa (year, month) -> total sueldo del mes
    meses_con_datos_map: dict[tuple[int, int], float] = {
        (int(mes.get("year") or 0), int(mes.get("month") or 0)): float(mes.get("total_sueldos") or 0)
        for mes in periodos
        if mes.get("year") and mes.get("month")
    }

    total_salarios = sum(meses_con_datos_map.values())
    aguinaldo = round(total_salarios / 12, 2)

    # Período considerado: desde el 01/01 del año de salida (o fecha_ingreso si es posterior)
    fecha_inicio_periodo: date | None = None
    if fecha_salida:
        inicio_anio = date(fecha_salida.year, 1, 1)
        fecha_inicio_periodo = (
            fecha_ingreso
            if fecha_ingreso and fecha_ingreso > inicio_anio
            else inicio_anio
        )

    # Lista de todos los meses del período (con o sin datos)
    detalle_meses: list[dict] = []
    if fecha_inicio_periodo and fecha_salida:
        cy, cm = fecha_inicio_periodo.year, fecha_inicio_periodo.month
        while (cy, cm) <= (fecha_salida.year, fecha_salida.month):
            monto = meses_con_datos_map.get((cy, cm))
            detalle_meses.append({
                "mes": f"{_mes_es(cm)} {cy}",
                "monto": monto,
                "tiene_datos": monto is not None,
            })
            cm += 1
            if cm > 12:
                cm = 1
                cy += 1

    meses_con_datos = len(meses_con_datos_map)
    meses_esperados = len(detalle_meses) if detalle_meses else (fecha_salida.month if fecha_salida else 12)
    meses_faltantes = sum(1 for m in detalle_meses if not m["tiene_datos"])

    periodo_texto = (
        f"{fecha_inicio_periodo.strftime('%d/%m/%Y')} al {fecha_salida.strftime('%d/%m/%Y')}"
        if fecha_inicio_periodo and fecha_salida
        else (" - ".join(m.get("mes_nombre", "") for m in periodos) or "Sin periodo")
    )

    return {
        "periodo": periodo_texto,
        "total_salarios": round(total_salarios, 2),
        "aguinaldo": aguinaldo,
        "meses_con_datos": meses_con_datos,
        "meses_esperados": meses_esperados,
        "meses_faltantes": meses_faltantes,
        "alerta_meses_incompletos": meses_faltantes > 0,
        "detalle_meses": detalle_meses,
        "fecha_inicio_periodo": fecha_inicio_periodo.strftime("%d/%m/%Y") if fecha_inicio_periodo else "",
        "fecha_fin_periodo": fecha_salida.strftime("%d/%m/%Y") if fecha_salida else "",
    }


def _dias_anuales_por_antiguedad(anios_cumplidos: int) -> int:
    if anios_cumplidos <= 5:
        return 12
    if anios_cumplidos <= 10:
        return 18
    return 30


def _calcular_vacaciones_por_liquidacion(
    empleado: Employee,
    registros: list[EmployeePayroll],
    fecha_salida: date | None,
    vacaciones_usadas: float,
) -> dict:
    fecha_ingreso = _obtener_fecha_ingreso_empleado(empleado, registros)
    if not fecha_ingreso or not fecha_salida:
        return {
            "fecha_ingreso": _formatear_fecha_iso(fecha_ingreso),
            "dias_generados": 0.0,
            "dias_usados": round(vacaciones_usadas, 2),
            "dias_pendientes": 0.0,
            "promedio_diario": 0.0,
            "promedio_mensual": 0.0,
            "monto": 0.0,
            "dias_antiguedad": 0,
            "periodo_promedio": "Sin periodo",
            "meses_detalle": [],
        }

    dias_antiguedad = max((fecha_salida - fecha_ingreso).days, 0)
    meses_periodo = _obtener_ultimos_periodos(registros, fecha_salida, cantidad=6)

    # Construir detalle mensual
    meses_detalle = [
        {"nombre": mes.get("mes_nombre", ""), "total": float(mes.get("total_sueldos") or 0)}
        for mes in meses_periodo
        if mes.get("mes_nombre")
    ]

    total_salarios = sum(m["total"] for m in meses_detalle)
    cant_meses = max(len(meses_detalle), 1)

    # Fórmula art. 92 inc. b CT: promedio mensual = total / meses; promedio diario = mensual / 30
    promedio_mensual = round(total_salarios / cant_meses, 2) if total_salarios else 0.0
    promedio_diario = round(promedio_mensual / 30, 2)

    periodo_promedio = " - ".join(m["nombre"] for m in meses_detalle)

    anios_completos = dias_antiguedad // 365
    dias_restantes = dias_antiguedad % 365
    dias_generados = 0.0
    for anio in range(1, anios_completos + 1):
        dias_generados += _dias_anuales_por_antiguedad(anio)
    if dias_restantes > 0:
        dias_generados += (dias_restantes / 365) * _dias_anuales_por_antiguedad(anios_completos + 1)

    dias_pendientes = round(max(dias_generados - vacaciones_usadas, 0.0), 2)
    monto = round(dias_pendientes * promedio_diario, 2)
    return {
        "fecha_ingreso": _formatear_fecha_iso(fecha_ingreso),
        "dias_generados": round(dias_generados, 2),
        "dias_usados": round(vacaciones_usadas, 2),
        "dias_pendientes": dias_pendientes,
        "promedio_diario": promedio_diario,
        "promedio_mensual": promedio_mensual,
        "monto": monto,
        "dias_antiguedad": dias_antiguedad,
        "periodo_promedio": periodo_promedio or "Sin periodo",
        "meses_detalle": meses_detalle,
    }


def _calcular_preaviso_e_indemnizacion(tipo_liquidacion: str, dias_antiguedad: int, promedio_diario: float) -> dict:
    if tipo_liquidacion != "despido-sin-causa":
        return {"preaviso": 0.0, "indemnizacion": 0.0, "preaviso_dias": 0, "indemnizacion_dias": 0}

    # Escala de preaviso: art. 87 CT (4 tramos)
    if dias_antiguedad <= 365:
        preaviso_dias = 30
    elif dias_antiguedad <= 5 * 365:
        preaviso_dias = 45
    elif dias_antiguedad <= 10 * 365:
        preaviso_dias = 60
    else:
        preaviso_dias = 90  # art. 87 CT: más de 10 años → 90 días

    # Indemnización: art. 91 CT — 15 días por año; fracción > 6 meses = 1 año
    anios_completos = dias_antiguedad // 365
    meses_fraccion = (dias_antiguedad % 365) // 30
    anios = anios_completos + (1 if meses_fraccion >= 6 else 0)
    anios = max(anios, 1)  # mínimo 1 año si ya superó los 6 meses (art. 91)
    indemnizacion_dias = 15 * anios
    return {
        "preaviso": round(preaviso_dias * promedio_diario, 2),
        "indemnizacion": round(indemnizacion_dias * promedio_diario, 2),
        "preaviso_dias": preaviso_dias,
        "indemnizacion_dias": indemnizacion_dias,
    }


def _dias_inclusive(fecha_inicio: date | None, fecha_fin: date | None) -> float:
    if not fecha_inicio or not fecha_fin:
        return 0.0
    if fecha_fin < fecha_inicio:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio
    return float((fecha_fin - fecha_inicio).days + 1)


def _calcular_timeline_antiguedad(
    fecha_ingreso: date | None,
    fecha_salida: date | None,
    meses_aguinaldo: int,
) -> dict:
    if not fecha_ingreso or not fecha_salida:
        return {
            "fecha_ingreso": "N/D",
            "fecha_salida": "N/D",
            "periodos_completos": 0,
            "periodo_proporcional": 0.0,
            "meses_aguinaldo": max(0, min(int(meses_aguinaldo or 0), 12)),
            "aguinaldo_porcentaje": 0.0,
            "donut_gradient": "conic-gradient(#3b82f6 0 0%, #334155 0 100%)",
        }

    if fecha_salida < fecha_ingreso:
        fecha_ingreso, fecha_salida = fecha_salida, fecha_ingreso

    periodos_completos = 0
    periodo_proporcional = 0.0

    cursor = date(fecha_ingreso.year, fecha_ingreso.month, 1)
    limite = date(fecha_salida.year, fecha_salida.month, 1)

    while cursor <= limite:
        dias_mes = calendar.monthrange(cursor.year, cursor.month)[1]
        inicio_mes = cursor
        fin_mes = date(cursor.year, cursor.month, dias_mes)

        tramo_inicio = max(fecha_ingreso, inicio_mes)
        tramo_fin = min(fecha_salida, fin_mes)

        if tramo_inicio <= tramo_fin:
            if tramo_inicio == inicio_mes and tramo_fin == fin_mes:
                periodos_completos += 1
            else:
                dias_trabajados_mes = (tramo_fin - tramo_inicio).days + 1
                periodo_proporcional += dias_trabajados_mes / dias_mes

        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)

    meses_aguinaldo = max(0, min(int(meses_aguinaldo or 0), 12))
    aguinaldo_porcentaje = round((meses_aguinaldo / 12) * 100, 2)
    donut_gradient = (
        f"conic-gradient(#3b82f6 0 {aguinaldo_porcentaje:.2f}%, #334155 {aguinaldo_porcentaje:.2f}% 100%)"
    )

    return {
        "fecha_ingreso": _formatear_fecha_iso(fecha_ingreso),
        "fecha_salida": fecha_salida.strftime("%d/%m/%Y"),
        "periodos_completos": periodos_completos,
        "periodo_proporcional": round(periodo_proporcional, 2),
        "meses_aguinaldo": meses_aguinaldo,
        "aguinaldo_porcentaje": aguinaldo_porcentaje,
        "donut_gradient": donut_gradient,
    }


def _calcular_promedio_laboral(registros: list[EmployeePayroll], empleado_id: int) -> dict:
    fechas = []
    total_salarios = 0.0
    dias_trabajados = 0

    for registro in registros:
      fecha_texto = getattr(registro, "fecha", "")
      if fecha_texto:
          fechas.append(str(fecha_texto))
      total_salarios += float(getattr(registro, "sueldo_final", 0) or 0)
      dias_trabajados += 1

    if not dias_trabajados:
        return {
            "periodo": "Sin periodo",
            "total_salarios": 0.0,
            "dias_trabajados": 0,
            "promedio_diario": 0.0,
            "promedio_mensual": 0.0,
        }

    promedio_diario = round(total_salarios / dias_trabajados, 2)
    promedio_mensual = round(promedio_diario * 30, 2)
    periodo = f"{min(fechas)} a {max(fechas)}" if fechas else "Sin periodo"

    return {
        "periodo": periodo,
        "total_salarios": round(total_salarios, 2),
        "dias_trabajados": dias_trabajados,
        "promedio_diario": promedio_diario,
        "promedio_mensual": promedio_mensual,
    }


def _calcular_liquidacion_resumen(tipo_liquidacion: str, promedio: dict, vacaciones_pendientes: float) -> dict:
    promedio_diario = float(promedio.get("promedio_diario") or 0.0)
    promedio_mensual = float(promedio.get("promedio_mensual") or 0.0)

    salario_pendiente = round(promedio_mensual, 2)
    aguinaldo = round(float(promedio.get("total_salarios") or 0.0) / 12, 2)
    vacaciones = round(vacaciones_pendientes * promedio_diario, 2)

    if tipo_liquidacion == "despido-sin-causa":
        preaviso = round(promedio_mensual, 2)
        indemnizacion = round(promedio_mensual, 2)
    elif tipo_liquidacion == "fin-de-contrato":
        preaviso = 0.0
        indemnizacion = 0.0
    else:
        preaviso = 0.0
        indemnizacion = 0.0

    total_liquidacion = round(salario_pendiente + aguinaldo + vacaciones + preaviso + indemnizacion, 2)
    return {
        "salario_pendiente": salario_pendiente,
        "aguinaldo": aguinaldo,
        "vacaciones": vacaciones,
        "preaviso": preaviso,
        "indemnizacion": indemnizacion,
        "total_liquidacion": total_liquidacion,
    }


def _label_tipo_liquidacion(tipo: str | None) -> str:
    labels = {
        "renuncia-voluntaria": "Renuncia voluntaria",
        "despido-sin-causa": "Despido sin causa",
        "despido-con-causa": "Despido con causa",
        "fin-de-contrato": "Fin de contrato",
    }
    return labels.get((tipo or "").strip(), "Tipo no definido")


def _detalle_calculo_liquidacion_historial(liq: Liquidacion) -> dict:
    salario_pendiente = float(liq.salario_pendiente or 0)
    aguinaldo = float(liq.aguinaldo or 0)
    vacaciones = float(liq.vacaciones or 0)
    preaviso = float(liq.preaviso or 0)
    indemnizacion = float(liq.indemnizacion or 0)
    total = float(liq.total_liquidacion or 0)

    return {
        "salario_pendiente": (
            f"Salario pendiente: Gs. {salario_pendiente:,.0f}. "
            "Corresponde al tramo trabajado hasta la fecha de salida seleccionada."
        ),
        "aguinaldo": (
            f"Aguinaldo proporcional: Gs. {aguinaldo:,.0f}. "
            "Se calcula prorrateando los periodos salariales acumulados sobre 12 meses."
        ),
        "vacaciones": (
            f"Vacaciones no gozadas: Gs. {vacaciones:,.0f}. "
            "Resultado de días pendientes por promedio diario vigente en la liquidación."
        ),
        "preaviso": (
            f"Preaviso: Gs. {preaviso:,.0f}. "
            + (
                "Aplica según tipo de liquidación y antigüedad." if preaviso > 0 else "No aplica para este caso."
            )
        ),
        "indemnizacion": (
            f"Indemnización: Gs. {indemnizacion:,.0f}. "
            + (
                "Aplica cuando corresponde legalmente por tipo de desvinculación." if indemnizacion > 0 else "No aplica para este caso."
            )
        ),
        "total": f"Total liquidación: Gs. {total:,.0f}. Suma final de todos los conceptos aplicables.",
    }


@app.route("/empleado/<int:empleado_id>/liquidaciones", methods=["GET", "POST"])
def liquidaciones(empleado_id):
    """Muestra la vista de liquidaciones para un empleado."""
    empleado = db.session.get(Employee, empleado_id)
    if not empleado:
        abort(404)

    registros = (
        EmployeePayroll.query.filter_by(employee_id=empleado_id)
        .order_by(text("employee_payroll.fecha ASC"))
        .all()
    )

    historial_mensual = [{"fecha": r.fecha} for r in registros if getattr(r, "fecha", None)]
    periodos_registrados = len(_agrupar_registros_por_mes_key(registros))
    promedio_actual = _calcular_promedio_laboral(registros, empleado_id)
    active_tab = (request.args.get("tab") or "nueva").strip().lower()
    if active_tab not in {"nueva", "historial", "exportar"}:
        active_tab = "nueva"

    historial_liquidaciones_raw = (
        Liquidacion.query.filter_by(empleado_id=empleado_id)
        .order_by(desc(Liquidacion.created_at))
        .all()
    )
    fecha_ingreso_empleado = _parse_fecha_form(getattr(empleado, "hire_date", None))
    historial_liquidaciones = []
    for item in historial_liquidaciones_raw:
        fecha_salida_item = _parse_fecha_form(getattr(item, "fecha_salida", None))
        detalle_texto = _detalle_calculo_liquidacion_historial(item)
        periodo_trabajado_str = "N/D"
        if fecha_salida_item:
            if fecha_ingreso_empleado:
                inicio_periodo = max(date(fecha_salida_item.year, 1, 1), fecha_ingreso_empleado)
            else:
                inicio_periodo = date(fecha_salida_item.year, 1, 1)
            periodo_trabajado_str = (
                f"{inicio_periodo.strftime('%d/%m/%Y')} - {fecha_salida_item.strftime('%d/%m/%Y')}"
            )
        historial_liquidaciones.append(
            {
                "id": item.id,
                "tipo": _label_tipo_liquidacion(item.tipo),
                "tipo_key": (item.tipo or "").strip(),
                "fecha_registro": item.created_at.strftime("%d/%m/%Y %H:%M") if item.created_at else "N/D",
                "fecha_salida": fecha_salida_item.strftime("%d/%m/%Y") if fecha_salida_item else "N/D",
                "periodo_trabajado": periodo_trabajado_str,
                "salario_pendiente": float(item.salario_pendiente or 0),
                "aguinaldo": float(item.aguinaldo or 0),
                "vacaciones": float(item.vacaciones or 0),
                "preaviso": float(item.preaviso or 0),
                "indemnizacion": float(item.indemnizacion or 0),
                "total_liquidacion": float(item.total_liquidacion or 0),
                "detalle_texto": detalle_texto,
            }
        )

    ultima_liquidacion = (
        Liquidacion.query.filter_by(empleado_id=empleado_id)
        .order_by(desc(Liquidacion.created_at))
        .first()
    )
    ultima_fecha_salida = _parse_fecha_form(getattr(ultima_liquidacion, "fecha_salida", None)) if ultima_liquidacion else None
    # Prioridad: último registro de nómina cargado > última liquidación > hoy
    ultima_fecha_nomina = None
    if registros:
        ultima_fecha_nomina = _parse_fecha_form(getattr(next(reversed(registros)), "fecha", None))
    fecha_salida_default = ultima_fecha_nomina or ultima_fecha_salida or date.today()
    try:
        estado_vacaciones = calcular_vacaciones_desde_historial(historial_mensual=historial_mensual)
    except ValueError:
        estado_vacaciones = {
            "fecha_inicio": "",
            "fecha_fin": "",
            "dias_trabajados": 0,
            "dias_vacaciones": 0.0,
            "mensaje": "Sin datos suficientes para calcular vacaciones",
        }

    vacaciones_usadas = float(empleado.vacation_used_days or 0.0)
    fecha_salida = fecha_salida_default
    ultimo_sueldo_pagado = False  # GET: default a no pagado (más seguro)
    salario_pendiente_data = _calcular_salario_pendiente_liquidacion(registros, fecha_salida, ultimo_sueldo_pagado)
    aguinaldo_data = _calcular_aguinaldo_proporcional(registros, fecha_salida, fecha_ingreso_empleado)
    vacaciones_data = _calcular_vacaciones_por_liquidacion(
        empleado=empleado,
        registros=registros,
        fecha_salida=fecha_salida,
        vacaciones_usadas=vacaciones_usadas,
    )
    preaviso_indemnizacion = _calcular_preaviso_e_indemnizacion(
        empleado.liquidation_type or "renuncia-voluntaria",
        vacaciones_data["dias_antiguedad"],
        vacaciones_data["promedio_diario"],
    )

    vacaciones_generadas = float(vacaciones_data["dias_generados"] or estado_vacaciones.get("dias_vacaciones") or 0.0)
    vacaciones_pendientes = float(vacaciones_data["dias_pendientes"] or 0.0)
    monto_vacaciones_pendientes = float(vacaciones_data["monto"] or 0.0)
    salario_pendiente = float(salario_pendiente_data["monto"] or 0.0)
    aguinaldo = float(aguinaldo_data["aguinaldo"] or 0.0)
    preaviso = float(preaviso_indemnizacion["preaviso"] or 0.0)
    indemnizacion = float(preaviso_indemnizacion["indemnizacion"] or 0.0)
    total_liquidacion = round(salario_pendiente + aguinaldo + monto_vacaciones_pendientes + preaviso + indemnizacion, 2)

    if request.method == "POST":
        tipo_liquidacion = (request.form.get("liq_tipo") or "renuncia-voluntaria").strip()
        fecha_salida = _parse_fecha_form(request.form.get("liq_salida")) or fecha_salida_default
        ultimo_sueldo_pagado = request.form.get("ultimo_sueldo_pagado", "no") == "si"

        salario_pendiente_data = _calcular_salario_pendiente_liquidacion(registros, fecha_salida, ultimo_sueldo_pagado)
        aguinaldo_data = _calcular_aguinaldo_proporcional(registros, fecha_salida, fecha_ingreso_empleado)

        vacaciones_usadas = float(request.form.get("vacaciones_usadas") or 0.0)
        fecha_desde = _parse_fecha_form(request.form.get("vacaciones_desde"))
        fecha_hasta = _parse_fecha_form(request.form.get("vacaciones_hasta"))
        if vacaciones_usadas <= 0 and fecha_desde and fecha_hasta:
            vacaciones_usadas = _dias_inclusive(fecha_desde, fecha_hasta)

        vacaciones_data = _calcular_vacaciones_por_liquidacion(
            empleado=empleado,
            registros=registros,
            fecha_salida=fecha_salida,
            vacaciones_usadas=vacaciones_usadas,
        )
        preaviso_indemnizacion = _calcular_preaviso_e_indemnizacion(
            tipo_liquidacion,
            vacaciones_data["dias_antiguedad"],
            vacaciones_data["promedio_diario"],
        )

        salario_pendiente = salario_pendiente_data["monto"]
        aguinaldo = aguinaldo_data["aguinaldo"]
        vacaciones = vacaciones_data["monto"]
        preaviso = preaviso_indemnizacion["preaviso"]
        indemnizacion = preaviso_indemnizacion["indemnizacion"]
        total_liquidacion = round(salario_pendiente + aguinaldo + vacaciones + preaviso + indemnizacion, 2)

        resumen = {
            "salario_pendiente": salario_pendiente,
            "aguinaldo": aguinaldo,
            "vacaciones": vacaciones,
            "preaviso": preaviso,
            "indemnizacion": indemnizacion,
            "total_liquidacion": total_liquidacion,
        }

        promedio = PromedioLaboral()
        promedio.empleado_id = empleado.id
        promedio.periodo = aguinaldo_data["periodo"] or promedio_actual["periodo"]
        promedio.total_salarios = aguinaldo_data["total_salarios"] or promedio_actual["total_salarios"]
        promedio.dias_trabajados = vacaciones_data["dias_antiguedad"] or promedio_actual["dias_trabajados"]
        promedio.promedio_diario = vacaciones_data["promedio_diario"] or promedio_actual["promedio_diario"]
        promedio.promedio_mensual = round((promedio.promedio_diario or 0) * 30, 2)

        empleado.liquidation_type = tipo_liquidacion
        empleado.vacation_generated_days = vacaciones_generadas
        empleado.vacation_used_days = vacaciones_usadas
        empleado.vacation_pending_days = vacaciones_pendientes
        empleado.vacation_used_from = request.form.get("vacaciones_desde") or None
        empleado.vacation_used_to = request.form.get("vacaciones_hasta") or None

        liquidacion = Liquidacion()
        liquidacion.empleado_id = empleado.id
        liquidacion.tipo = tipo_liquidacion
        liquidacion.fecha_salida = _formatear_fecha_iso(fecha_salida)
        liquidacion.salario_pendiente = resumen["salario_pendiente"]
        liquidacion.aguinaldo = resumen["aguinaldo"]
        liquidacion.vacaciones = resumen["vacaciones"]
        liquidacion.preaviso = resumen["preaviso"]
        liquidacion.indemnizacion = resumen["indemnizacion"]
        liquidacion.total_liquidacion = resumen["total_liquidacion"]
        db.session.add(promedio)
        db.session.add(liquidacion)
        db.session.commit()

        flash("La liquidación se guardó con las vacaciones actualizadas.", "success")
        return redirect(url_for("liquidaciones", empleado_id=empleado_id, tab="historial"))

    ultimo_promedio = (
        PromedioLaboral.query.filter_by(empleado_id=empleado_id)
        .order_by(desc(PromedioLaboral.created_at))
        .first()
    )

    detalle_vacaciones = (
        f"Vacaciones generadas: {vacaciones_generadas:.2f} días · "
        f"Vacaciones utilizadas: {vacaciones_usadas:.2f} días · "
        f"Vacaciones pendientes: {vacaciones_pendientes:.2f} días"
    )
    _sp = salario_pendiente_data
    _lineas_sp: list[str] = []
    if _sp.get("fecha_inicio") and _sp.get("fecha_fin"):
        _lineas_sp.append(f"Período pendiente: {_sp['fecha_inicio']} al {_sp['fecha_fin']}")
    _lineas_sp.append(f"Horas normales: <strong>{_fmt_hhmm(_sp.get('horas_normales', 0.0))}</strong>")
    _lineas_sp.append(f"Horas especiales: <strong>{_fmt_hhmm(_sp.get('horas_especiales', 0.0))}</strong>")
    if _sp.get("horas_feriado", 0.0) > 0:
        _lineas_sp.append(f"Horas feriado: <strong>{_fmt_hhmm(_sp.get('horas_feriado', 0.0))}</strong> (×2)")
    _lineas_sp.append(f"Valor hora: <strong>Gs. {_sp.get('valor_hora', 0.0):,.0f}</strong>")
    _lineas_sp.append(f"Total: <strong>Gs. {salario_pendiente:,.0f}</strong>")
    detalle_salario_pendiente = "\n".join(_lineas_sp)
    meses_aguinaldo = len(_obtener_ultimos_periodos(registros, fecha_salida_default, 12))
    # Desglose mes a mes del aguinaldo (auditable)
    _lineas_ag = [f"Período: {aguinaldo_data['fecha_inicio_periodo']} al {aguinaldo_data['fecha_fin_periodo']}"]
    for _m in aguinaldo_data.get("detalle_meses", []):
        if _m["tiene_datos"]:
            _lineas_ag.append(f"{_m['mes']}: <strong>Gs. {_m['monto']:,.0f}</strong>")
        else:
            _lineas_ag.append(f"{_m['mes']}: <em>sin registros</em>")
    _lineas_ag.append(f"Total salarios: <strong>Gs. {aguinaldo_data['total_salarios']:,.0f}</strong>")
    _lineas_ag.append(
        f"Aguinaldo = Gs. {aguinaldo_data['total_salarios']:,.0f} / 12 = "
        f"<strong>Gs. {aguinaldo_data['aguinaldo']:,.0f}</strong>"
    )
    detalle_aguinaldo = "\n".join(_lineas_ag)
    aguinaldo_alerta = aguinaldo_data.get("alerta_meses_incompletos", False)
    aguinaldo_meses_con_datos = aguinaldo_data.get("meses_con_datos", meses_aguinaldo)
    aguinaldo_meses_esperados = aguinaldo_data.get("meses_esperados", 12)
    aguinaldo_meses_faltantes = aguinaldo_data.get("meses_faltantes", 0)
    detalle_vacaciones = (
        f"{detalle_vacaciones} · {vacaciones_data['dias_pendientes']:.2f} días × Gs. {vacaciones_data['promedio_diario']:,.0f}"
    )
    detalle_preaviso = (
        f"{preaviso_indemnizacion['preaviso_dias']} días × Gs. {vacaciones_data['promedio_diario']:,.0f}"
        if preaviso > 0
        else "No aplica"
    )
    detalle_indemnizacion = (
        f"{preaviso_indemnizacion['indemnizacion_dias']} días × Gs. {vacaciones_data['promedio_diario']:,.0f}"
        if indemnizacion > 0
        else "No aplica"
    )
    timeline_antiguedad = _calcular_timeline_antiguedad(
        _parse_fecha_form(vacaciones_data.get("fecha_ingreso")),
        fecha_salida_default,
        meses_aguinaldo,
    )

    composicion_base = [
        {
            "label": "Salario pendiente",
            "amount": float(salario_pendiente or 0),
            "color": "#4f83ff",
            "dot_class": "dot-blue",
        },
        {
            "label": "Aguinaldo proporcional",
            "amount": float(aguinaldo or 0),
            "color": "#22c55e",
            "dot_class": "dot-green",
        },
        {
            "label": "Vacaciones no gozadas",
            "amount": float(monto_vacaciones_pendientes or 0),
            "color": "#facc15",
            "dot_class": "dot-yellow",
        },
    ]
    if float(indemnizacion or 0) > 0:
        composicion_base.append(
            {
                "label": "Indemnización",
                "amount": float(indemnizacion or 0),
                "color": "#fb923c",
                "dot_class": "dot-orange",
            }
        )
    if float(preaviso or 0) > 0:
        composicion_base.append(
            {
                "label": "Preaviso",
                "amount": float(preaviso or 0),
                "color": "#ef4444",
                "dot_class": "dot-red",
            }
        )

    composicion_liquidacion = []
    composicion_chart_gradient = "conic-gradient(#64748b 0 100%)"
    composicion_dominante_label = "Sin datos"
    composicion_dominante_porcentaje = 0.0
    total_composicion = float(total_liquidacion or 0)
    if total_composicion > 0:
        acumulado = 0.0
        segmentos = []
        for concepto in composicion_base:
            porcentaje = round((concepto["amount"] / total_composicion) * 100, 2)
            composicion_liquidacion.append(
                {
                    **concepto,
                    "percentage": porcentaje,
                }
            )
            inicio = acumulado
            acumulado = min(acumulado + porcentaje, 100.0)
            segmentos.append(f"{concepto['color']} {inicio:.2f}% {acumulado:.2f}%")

        if segmentos:
            composicion_chart_gradient = f"conic-gradient({', '.join(segmentos)})"
        dominante = max(composicion_liquidacion, key=lambda item: float(item.get("percentage") or 0))
        composicion_dominante_label = str(dominante.get("label") or "Sin datos")
        composicion_dominante_porcentaje = float(dominante.get("percentage") or 0)
    else:
        for concepto in composicion_base:
            composicion_liquidacion.append(
                {
                    **concepto,
                    "percentage": 0.0,
                }
            )

    return render_template(
        "liquidaciones.html",
        empleado=empleado,
        mensaje=estado_vacaciones["mensaje"],
        detalle_vacaciones=detalle_vacaciones,
        vacaciones_generadas=round(vacaciones_generadas, 2),
        vacaciones_usadas=round(vacaciones_usadas, 2),
        vacaciones_pendientes=round(vacaciones_pendientes, 2),
        monto_vacaciones_pendientes=monto_vacaciones_pendientes,
        vacaciones_desde=empleado.vacation_used_from or "",
        vacaciones_hasta=empleado.vacation_used_to or "",
        selected_liquidation_type=empleado.liquidation_type or "renuncia-voluntaria",
        promedio_laboral=ultimo_promedio,
        ultima_liquidacion=ultima_liquidacion,
        periodos_registrados=periodos_registrados,
        fecha_ingreso=vacaciones_data["fecha_ingreso"],
        vacaciones_dias_antiguedad=vacaciones_data["dias_antiguedad"],
        salario_pendiente=salario_pendiente,
        aguinaldo=aguinaldo,
        preaviso=preaviso,
        indemnizacion=indemnizacion,
        total_liquidacion=total_liquidacion,
        fecha_salida_default=_formatear_fecha_iso(fecha_salida_default),
        promedio_diario_vacaciones=vacaciones_data["promedio_diario"],
        promedio_mensual_vacaciones=vacaciones_data["promedio_mensual"],
        vacaciones_meses_promedio=vacaciones_data["meses_detalle"],
        vacaciones_periodo_promedio=vacaciones_data["periodo_promedio"],
        detalle_salario_pendiente=detalle_salario_pendiente,
        detalle_aguinaldo=detalle_aguinaldo,
        detalle_preaviso=detalle_preaviso,
        detalle_indemnizacion=detalle_indemnizacion,
        composicion_liquidacion=composicion_liquidacion,
        composicion_chart_gradient=composicion_chart_gradient,
        composicion_dominante_label=composicion_dominante_label,
        composicion_dominante_porcentaje=composicion_dominante_porcentaje,
        timeline_antiguedad=timeline_antiguedad,
        active_tab=active_tab,
        historial_liquidaciones=historial_liquidaciones,
        aguinaldo_alerta=aguinaldo_alerta,
        aguinaldo_meses_con_datos=aguinaldo_meses_con_datos,
        aguinaldo_meses_esperados=aguinaldo_meses_esperados,
        aguinaldo_meses_faltantes=aguinaldo_meses_faltantes,
        ultimo_sueldo_pagado=ultimo_sueldo_pagado,
        salario_modo=salario_pendiente_data.get("modo", ""),
    )


@app.route("/empleado/<int:empleado_id>/descargar_mes/<int:year>/<int:month>")
def descargar_mes(empleado_id, year, month):
    empleado = db.session.get(Employee, empleado_id)
    if not empleado:
        abort(404)

    registros = (
        EmployeePayroll.query.filter_by(employee_id=empleado_id)
        .order_by(desc(EmployeePayroll.created_at))
        .all()
    )
    registros_mes = _filtrar_registros_por_mes(registros, year, month)
    if not registros_mes:
        abort(404)

    formato = request.args.get("format", "xlsx").lower()
    mes_nombre = f"{MESES_ES[month - 1]} {year}"
    base_nombre = _sanear_nombre_para_archivo(f"{empleado.nombre}_{mes_nombre}")

    if formato == "pdf":
        contenido = _crear_pdf_desde_registros(registros_mes, empleado.nombre, mes_nombre)
        mimetype = "application/pdf"
        nombre_archivo = f"{base_nombre}.pdf"
    else:
        contenido = _crear_excel_desde_registros(registros_mes, empleado.nombre, mes_nombre)
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        nombre_archivo = f"{base_nombre}.xlsx"

    return send_file(
        contenido,
        mimetype=mimetype,
        as_attachment=True,
        download_name=nombre_archivo,
    )


@app.route("/empleado/<int:empleado_id>/eliminar", methods=["POST"])
def eliminar_empleado(empleado_id):
    empleado = db.session.get(Employee, empleado_id)
    if not empleado:
        flash("El empleado no existe o ya fue eliminado.", "error")
        return redirect(url_for("index"))

    db.session.delete(empleado)
    db.session.commit()
    flash(f"Empleado {empleado.nombre} y sus datos asociados fueron eliminados.", "success")
    return redirect(url_for("index"))


@app.route("/procesar", methods=["POST"])
def procesar():
    try:
        valor_por_hora = float(request.form.get("valor_por_hora", VALOR_POR_HORA_DEFAULT))
    except (TypeError, ValueError):
        valor_por_hora = VALOR_POR_HORA_DEFAULT

    employee_id = request.form.get("employee_id", type=int)
    if not employee_id:
        flash("Selecciona un empleado para asociar la carga de asistencia.", "error")
        return redirect(url_for("index"))

    empleado = db.session.get(Employee, employee_id)
    if not empleado:
        flash("El empleado seleccionado no existe.", "error")
        return redirect(url_for("index"))

    feriados = _parse_feriados(request.form)
    tipo_archivo = request.form.get("file_type", "excel")

    # --- Leer archivo(s) -------------------------------------------------- #
    try:
        if tipo_archivo == "pdf":
            archivos = request.files.getlist("pdf_files") or request.files.getlist("pdf_files[]")
            archivos = [a for a in archivos if a and a.filename]
            if not archivos:
                flash("No se cargó ningún archivo PDF.", "error")
                return redirect(url_for("index"))
            if len(archivos) > 2:
                flash("Máximo 2 archivos PDF permitidos (uno por quincena).", "error")
                return redirect(url_for("index"))

            dataframes, nombres = [], []
            for archivo in archivos:
                df_temp = procesar_pdf_a_dataframe(archivo)
                if df_temp.empty:
                    continue

                for col_hora in ("Entrada", "Salida"):
                    if col_hora in df_temp.columns:
                        df_temp[col_hora] = df_temp[col_hora].apply(_normalizar_hora_pdf)

                es_valido, _ = validar_datos_pdf(df_temp)
                if not es_valido:
                    flash(
                        f"El PDF {archivo.filename} tiene filas con formato irregular; se intentará procesar igual.",
                        "warning",
                    )

                dataframes.append(df_temp)
                nombres.append(archivo.filename)

            if not dataframes:
                flash(
                    "No se pudieron extraer datos de los PDF. Verifica el contenido.",
                    "error",
                )
                return redirect(url_for("index"))

            df = pd.concat(dataframes, ignore_index=True)
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
            df = df.sort_values(["Fecha", "Empleado"]).reset_index(drop=True)
            source_name = (
                nombres[0] if len(nombres) == 1 else f"combinado_{len(nombres)}_pdfs"
            )
        else:
            archivo = request.files.get("excel_file")
            if not archivo or not archivo.filename:
                flash("No se cargó ningún archivo Excel.", "error")
                return redirect(url_for("index"))
            df = pd.read_excel(archivo)
            es_valido, faltantes = validar_archivo_excel(df)
            if not es_valido:
                flash(
                    "El Excel no contiene las columnas: " + ", ".join(faltantes),
                    "error",
                )
                return redirect(url_for("index"))
            source_name = archivo.filename
    except Exception as e:  # noqa: BLE001
        flash(f"Error al procesar el archivo: {e}", "error")
        return redirect(url_for("index"))

    # --- Filtrar sin asistencia y detectar problemas --------------------- #
    df_con_asistencia, df_sin_asistencia = filtrar_registros_sin_asistencia(df)
    df = df_con_asistencia.reset_index(drop=True)

    if tipo_archivo == "pdf":
        if df.empty:
            flash(
                f"No se encontraron registros de asistencia para {empleado.nombre} en los PDF.",
                "error",
            )
            return redirect(url_for("index", empleado_id=employee_id))
    else:
        coincidencia = df["Empleado"].astype(str).apply(lambda x: _empleado_coincide(x, empleado))
        df = df[coincidencia].copy()
        if df.empty:
            flash(
                f"No se encontraron registros de asistencia para {empleado.nombre} en el archivo.",
                "error",
            )
            return redirect(url_for("index", empleado_id=employee_id))

    df.loc[:, "Empleado"] = empleado.nombre

    df_incompletos = detectar_registros_incompletos(df)
    df_ambiguos = detectar_horarios_ambiguos(df)

    ips_value = (request.form.get("seguro_ips", "No") or "No").strip().lower()
    ips_enabled = ips_value in ("sí", "si", "s", "yes", "true")

    config = {
        "valor_por_hora": valor_por_hora,
        "feriados": list(feriados),
        "source_name": source_name,
        "source_type": tipo_archivo,
        "excluidos": len(df_sin_asistencia),
        "employee_id": employee_id,
        "ips_enabled": ips_enabled,
    }

    meses_df = _obtener_meses_df(df)
    meses_conflictivos, registros_conflictivos = _obtener_conflictos_mes_empleado(
        employee_id,
        meses_df,
    )
    if meses_conflictivos:
        config["meses_conflictivos"] = meses_conflictivos
        token_reemplazo = _guardar_pendiente(df, config)
        session["replace_pending_token"] = token_reemplazo
        return _render_confirmacion_reemplazo(
            empleado,
            token_reemplazo,
            meses_conflictivos,
            registros_conflictivos,
        )

    if not df_incompletos.empty or not df_ambiguos.empty:
        token = _guardar_pendiente(df, config)
        session["pending_token"] = token
        return render_template(
            "corrections.html",
            incompletos=_df_incompletos_a_lista(df_incompletos),
            ambiguos=_df_ambiguos_a_lista(df_ambiguos),
            excluidos=len(df_sin_asistencia),
        )

    run_id, _ = _calcular_y_guardar(df, config, empleado)
    return redirect(url_for("resultado", run_id=run_id))


@app.route("/correcciones", methods=["POST"])
def correcciones():
    token = session.get("pending_token")
    pendiente = _cargar_pendiente(token)
    if not pendiente:
        flash("La sesión de corrección expiró. Vuelve a subir el archivo.", "error")
        return redirect(url_for("index"))

    df = pendiente["df"].copy()
    config = pendiente["config"]
    employee_id = config.get("employee_id")
    empleado = db.session.get(Employee, employee_id)
    if not empleado:
        flash("El empleado asociado a la corrección ya no existe.", "error")
        return redirect(url_for("index"))

    # Registros incompletos: el admin indica si la marca fue Entrada o Salida
    # y aporta el horario faltante.
    df = _aplicar_correcciones_incompletos(df, pendiente["df"], request.form)

    # Horarios ambiguos: intercambiar entrada <-> salida
    for key in request.form:
        if key.startswith("intercambiar_"):
            _, idx_str = key.split("_", 1)
            idx = int(idx_str)
            entrada = df.at[idx, "Entrada"]
            df.at[idx, "Entrada"] = df.at[idx, "Salida"]
            df.at[idx, "Salida"] = entrada

    _eliminar_pendiente(token)
    session.pop("pending_token", None)

    meses_df = _obtener_meses_df(df)
    meses_conflictivos, registros_conflictivos = _obtener_conflictos_mes_empleado(
        employee_id,
        meses_df,
    )
    if meses_conflictivos:
        config["meses_conflictivos"] = meses_conflictivos
        token_reemplazo = _guardar_pendiente(df, config)
        session["replace_pending_token"] = token_reemplazo
        return _render_confirmacion_reemplazo(
            empleado,
            token_reemplazo,
            meses_conflictivos,
            registros_conflictivos,
        )

    run_id, _ = _calcular_y_guardar(df, config, empleado)
    return redirect(url_for("resultado", run_id=run_id))


@app.route("/actualizar-mes", methods=["POST"])
def actualizar_mes():
    token = request.form.get("pending_token") or session.get("replace_pending_token")
    pendiente = _cargar_pendiente(token)
    if not pendiente:
        flash("La solicitud de actualización expiró. Vuelve a cargar el archivo.", "error")
        return redirect(url_for("index"))

    df = pendiente["df"].copy()
    config = pendiente["config"]
    employee_id = config.get("employee_id")
    empleado = db.session.get(Employee, employee_id)
    if not empleado:
        flash("El empleado asociado a la actualización ya no existe.", "error")
        return redirect(url_for("index"))

    meses_df = _obtener_meses_df(df)
    _reemplazar_meses_existentes(employee_id, meses_df)

    _eliminar_pendiente(token)
    session.pop("replace_pending_token", None)

    run_id, _ = _calcular_y_guardar(df, config, empleado)
    flash("Se actualizó el mes y se reemplazaron los registros anteriores.", "success")
    return redirect(url_for("resultado", run_id=run_id))


def _aplicar_correcciones_incompletos(df, df_original, form):
    """
    Completa registros incompletos respetando el horario realmente registrado.
    El administrador indica si la marca registrada fue Entrada o Salida y aporta
    el horario faltante.
    """
    for key in list(form.keys()):
        if not key.startswith("tipo_"):
            continue
        idx = int(key.split("_", 1)[1])
        tipo = form.get(key)
        hora_faltante = (form.get(f"hora_{idx}") or "").strip()
        if not hora_faltante:
            continue

        ent_orig = str(df_original.at[idx, "Entrada"]).strip()
        sal_orig = str(df_original.at[idx, "Salida"]).strip()
        registrado = sal_orig if ent_orig in ("", "nan", "0:00", "00:00") else ent_orig

        if tipo == "Entrada":
            df.at[idx, "Entrada"] = registrado
            df.at[idx, "Salida"] = hora_faltante
        else:
            df.at[idx, "Entrada"] = hora_faltante
            df.at[idx, "Salida"] = registrado

    return df


@app.route("/resultado/<int:run_id>")
def resultado(run_id):
    run = db.session.get(CalculationRun, run_id)
    if not run:
        abort(404)
    _recalcular_run_legado(run)
    records = cast(list[EmployeeRecord], list(getattr(run, "records", []) or []))
    return render_template("results.html", run=run, records=records)


@app.route("/descargar/<int:run_id>")
def descargar(run_id):
    run = db.session.get(CalculationRun, run_id)
    if not run:
        abort(404)

    _recalcular_run_legado(run)
    records = cast(list[EmployeeRecord], list(getattr(run, "records", []) or []))
    filas = []
    for r in records:
        filas.append(
            {
                "Empleado": r.empleado,
                "Fecha": r.fecha,
                "Entrada": r.entrada,
                "Salida": r.salida,
                "Feriado": r.feriado,
                "Horas Trabajadas (h:mm)": r.horas_trabajadas,
                "Horas Normales": horas_a_horasminutos(r.horas_normales),
                "Horas Especiales": horas_a_horasminutos(r.horas_especiales),
                "Descuento Inventario": r.descuento_inventario,
                "Descuento Caja": r.descuento_caja,
                "Retiro": r.retiro,
                "Sueldo Final": r.sueldo_final,
            }
        )

    df = pd.DataFrame(filas)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sueldos")
    output.seek(0)

    base = (run.source_name or "sueldos").replace(".pdf", "").replace(".xlsx", "")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{base}_calculado.xlsx",
    )


@app.route("/plantilla")
def plantilla():
    ruta = os.path.join(BASE_DIR, "plantilla_sueldos_feriados_dias.xlsx")
    if not os.path.exists(ruta):
        abort(404)
    return send_file(
        ruta,
        as_attachment=True,
        download_name="plantilla_sueldo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/eliminar/<int:run_id>", methods=["POST"])
def eliminar(run_id):
    run = db.session.get(CalculationRun, run_id)
    if run:
        db.session.delete(run)
        db.session.commit()
        flash("Cálculo eliminado del historial.", "success")
    return redirect(url_for("index"))


def _ensure_missing_employee_columns():
    engine = db.engine
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    def ensure_column(table_name: str, column_name: str, ddl: str) -> None:
        if table_name not in table_names:
            return
        current_columns = {col["name"] for col in inspector.get_columns(table_name)}
        if column_name in current_columns:
            return
        with engine.begin() as connection:
            connection.execute(text(ddl))
        app.logger.info("Added missing %s column to %s table", column_name, table_name)

    ensure_column(
        "calculation_runs",
        "employee_id",
        "ALTER TABLE calculation_runs ADD COLUMN employee_id INTEGER",
    )
    ensure_column(
        "calculation_runs",
        "seguro_ips",
        "ALTER TABLE calculation_runs ADD COLUMN seguro_ips VARCHAR(5) NOT NULL DEFAULT 'No'",
    )
    ensure_column(
        "calculation_runs",
        "total_salario_bruto",
        "ALTER TABLE calculation_runs ADD COLUMN total_salario_bruto FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_descuento_ips",
        "ALTER TABLE calculation_runs ADD COLUMN total_descuento_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_aporte_empleador_ips",
        "ALTER TABLE calculation_runs ADD COLUMN total_aporte_empleador_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_ips",
        "ALTER TABLE calculation_runs ADD COLUMN total_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_salario_neto_ips",
        "ALTER TABLE calculation_runs ADD COLUMN total_salario_neto_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_monto_horas_normales",
        "ALTER TABLE calculation_runs ADD COLUMN total_monto_horas_normales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_monto_horas_especiales",
        "ALTER TABLE calculation_runs ADD COLUMN total_monto_horas_especiales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_monto_feriados",
        "ALTER TABLE calculation_runs ADD COLUMN total_monto_feriados FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "calculation_runs",
        "total_bonificacion",
        "ALTER TABLE calculation_runs ADD COLUMN total_bonificacion FLOAT NOT NULL DEFAULT 0",
    )

    ensure_column(
        "employees",
        "hire_date",
        "ALTER TABLE employees ADD COLUMN hire_date VARCHAR(20)",
    )
    ensure_column(
        "employees",
        "liquidation_type",
        "ALTER TABLE employees ADD COLUMN liquidation_type VARCHAR(50)",
    )
    ensure_column(
        "employees",
        "vacation_generated_days",
        "ALTER TABLE employees ADD COLUMN vacation_generated_days FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employees",
        "vacation_used_days",
        "ALTER TABLE employees ADD COLUMN vacation_used_days FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employees",
        "vacation_pending_days",
        "ALTER TABLE employees ADD COLUMN vacation_pending_days FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employees",
        "vacation_used_from",
        "ALTER TABLE employees ADD COLUMN vacation_used_from VARCHAR(20)",
    )
    ensure_column(
        "employees",
        "vacation_used_to",
        "ALTER TABLE employees ADD COLUMN vacation_used_to VARCHAR(20)",
    )

    ensure_column(
        "employee_records",
        "employee_id",
        "ALTER TABLE employee_records ADD COLUMN employee_id INTEGER",
    )
    ensure_column(
        "employee_records",
        "descuento_ips",
        "ALTER TABLE employee_records ADD COLUMN descuento_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_records",
        "monto_horas_normales",
        "ALTER TABLE employee_records ADD COLUMN monto_horas_normales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_records",
        "monto_horas_especiales",
        "ALTER TABLE employee_records ADD COLUMN monto_horas_especiales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_records",
        "monto_feriado",
        "ALTER TABLE employee_records ADD COLUMN monto_feriado FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_records",
        "bonificacion",
        "ALTER TABLE employee_records ADD COLUMN bonificacion FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_records",
        "sueldo_bruto",
        "ALTER TABLE employee_records ADD COLUMN sueldo_bruto FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "descuento_ips",
        "ALTER TABLE employee_payroll ADD COLUMN descuento_ips FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "monto_horas_normales",
        "ALTER TABLE employee_payroll ADD COLUMN monto_horas_normales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "monto_horas_especiales",
        "ALTER TABLE employee_payroll ADD COLUMN monto_horas_especiales FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "monto_feriado",
        "ALTER TABLE employee_payroll ADD COLUMN monto_feriado FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "bonificacion",
        "ALTER TABLE employee_payroll ADD COLUMN bonificacion FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "sueldo_bruto",
        "ALTER TABLE employee_payroll ADD COLUMN sueldo_bruto FLOAT NOT NULL DEFAULT 0",
    )
    ensure_column(
        "employee_payroll",
        "valor_hora_utilizado",
        "ALTER TABLE employee_payroll ADD COLUMN valor_hora_utilizado FLOAT",
    )


with app.app_context():
    db.create_all()
    _ensure_missing_employee_columns()

    # Crear empleados iniciales solo si la tabla de empleados está vacía.
    # Esto evita que un empleado eliminado vuelva a aparecer al reiniciar la aplicación.
    if Employee.query.count() == 0:
        empleados_iniciales = ["Paz", "Yanina Gomez"]
        for nombre in empleados_iniciales:
            empleado = Employee(nombre=nombre)  # type: ignore[call-arg]
            db.session.add(empleado)
        db.session.commit()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
